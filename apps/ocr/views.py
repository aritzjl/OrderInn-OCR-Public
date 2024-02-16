from django.shortcuts import render
from django.http import HttpResponse
import easyocr
from io import BytesIO
from PIL import Image
from openai import OpenAI
from django.shortcuts import render
from django.http import HttpResponse
import easyocr
import pandas as pd
from django.contrib.auth.decorators import login_required
from io import BytesIO
import re
import os
from google.cloud import vision
from google.cloud.vision_v1 import types
from openpyxl import Workbook
import csv
from io import StringIO
from PIL import Image
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openai import OpenAI


def text_to_excel(text):
    print(text)
    text=text.replace(',','')
    client=OpenAI(api_key="")
    messages=[]
    mensaje="""
        Eres una herramienta que sirve para clasificar productos.
        Como veras en los ejemplos, si hay diferentes variantes de un mismo producto, deberas clasificarlo como productos a parte.
        Importante, debes poner todos los campos aunque esten vacios, y Original_Price SIEMPRE debe ser un numero
        RECUERDA, si algun campo no tiene info, como la descripcion, o el precio etc. Pon N/A como valor predeterminado, gracias.
        IMPORTANTE, si un item no tiene precio, ponle de precio 0.
       
    """
    message={"role": "system", "content": mensaje}
    messages.append(message)
    

    mensaje=open('mensaje1.txt','r').read()
    message={"role": "system", "content": mensaje}
    messages.append(message)
    
    mensaje=open('mensaje2.txt','r').read()
    message={"role": "assistant", "content": mensaje}
    messages.append(message)
    

    mensaje=open('mensaje3.txt','r').read()
    message={"role": "system", "content": mensaje}
    messages.append(message)
    
    mensaje=open('mensaje4.txt','r').read()
    message={"role": "assistant", "content": mensaje}
    messages.append(message)
    
    
    mensaje='Perfecto, ahora haz lo mismo, pero con este nuevo menu: ' +text
    message={"role": "user", "content": mensaje}
    messages.append(message)



    completion = client.chat.completions.create(
    model="gpt-3.5-turbo-16k",
    messages=messages
    )

    respuestaCSV=completion.choices[0].message.content
    print('===================================')
    print(respuestaCSV)
    print('===================================')
    return respuestaCSV

"""def process_image(image):
    # Procesar la imagen usando EasyOCR
    reader = easyocr.Reader(['en', 'es'])  # 'en' para inglés, puedes cambiarlo según el idioma
    result = reader.readtext(image)
    processed_text = '\n'.join([text[1] for text in result])
    return processed_text"""
    
    
def process_image(image):    
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r'light-router-385414-9c28f10f787b.json'
    client = vision.ImageAnnotatorClient()
    image_content = types.Image(content=image)
    
    response = client.text_detection(image=image_content)
    texts = response.text_annotations
    
    data = {'locale': [], 'description': []}
    
    for text in texts:
        data['locale'].append(text.locale)
        data['description'].append(text.description)
    
    df = pd.DataFrame(data)
    print(df['description'][0])
    return df['description'][0]

@login_required
def ocr(request):
    if request.method == 'POST' and request.FILES.get('image'):
        uploaded_image = request.FILES['image']
        
        # Convertir la imagen a un formato compatible
        img = Image.open(uploaded_image)
        img_byte_array = BytesIO()
        img.save(img_byte_array, format=img.format)
        
        processed_text = process_image(img_byte_array.getvalue())
# Procesar la imagen usando GPT Vision
        vision_result = text_to_excel(processed_text)

        # Convertir el texto en formato CSV a una lista de filas CSV, filtrando líneas vacías
        csv_rows = [row for row in vision_result.split('\n') if row.strip()]

        # Convertir el texto en formato CSV a una lista de filas CSV, filtrando líneas vacías
        csv_data = [row.split(':') for row in csv_rows]

        # Crear un archivo CSV temporal en memoria usando StringIO
        temp_csv = StringIO()
        writer = csv.writer(temp_csv)
        writer.writerows(csv_data)

        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active

        # Agregar encabezados y dar formato
        columnas = ["Category_Name", "Item_Name", "Item_Description", "Original_Price"]
        ws.append(columnas)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Leer el archivo CSV temporal y escribir los datos en el libro de trabajo de Excel
        temp_csv.seek(0)  # Mover el puntero al inicio del archivo en StringIO
        reader = csv.reader(temp_csv)
        data=[]
        for row in reader:
            print(row)
            # Reemplazar el separador alternativo de nuevo a comas en las descripciones
            if len(row)>1:
                row[1] = row[1].replace(' | ', ', ')
                data.append(row)
                
                
        # Crear un libro de trabajo de Excel
        workbook = Workbook()
        sheet = workbook.active

        # Agregar encabezados y dar formato
        columnas = ["Category_Name", "Item_Name", "Item_Description", "Original_Price"]
        sheet.append(columnas)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        finalrows=[]
        contador=0
        while contador<=len(data):
            try:
                categoria=data[contador][1]
                nombre=data[contador+1][1]
                descripcion=data[contador+2][1]
                precio=data[contador+3][1]
                precio=re.sub(r"[^\d.,]", "", precio)
                newRow=[categoria,nombre,descripcion,precio]
                finalrows.append(newRow)
                
            except:
                pass
            contador+=4
            
        for row in finalrows:
            sheet.append(row)
        

        # Guardar el libro de trabajo de Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="result.xlsx"'
        workbook.save(response)


        return response
    else:
        return render(request, 'upload_image.html')
